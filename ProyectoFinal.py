import pandas as pd
from openpyxl import Workbook, load_workbook
import numpy as np
excel_file = "C:/Users/defp_/OneDrive/Documentos/inventario.xlsx"

# inventario = pd.read_excel(io="C:/Users/defp_/OneDrive/Documentos/inventario.xlsx", sheet_name="Hoja2", engine="openpyxl",)

def inventory(excel):   
    inventario = pd.read_excel(excel, sheet_name="Hoja2")
    return inventario
inventario = inventory(excel_file)


def costo_promedio(inventario):
    costo_total = inventario['Stock Actual'] * inventario['Costo Unitario']
    total_stock = np.sum(inventario['Stock Actual'])
    promedio = costo_total / total_stock
    return promedio, total_stock, costo_total
costo = costo_promedio(inventario)

def add_inventory(name, stock, price):
    workbook = load_workbook('C:/Users/defp_/OneDrive/Documentos/inventario.xlsx')
    hoja2 = workbook['Hoja2']
    last_row = hoja2.max_row
    new_row = [name, stock, price]

    hoja2.append(new_row)

    workbook.save('C:/Users/defp_/OneDrive/Documentos/inventario.xlsx')
    
# add_inventory("Llave de 32 milimetros", 12, 100)


def update_file(inventario, costo):
    inventario['Costo total'] = costo[2]
    inventario['Costo promedio'] = costo[0]
    return inventario
inventory_updated = update_file(inventario, costo)

def rewrite_file(inventario, file):
    inventario.to_excel(file , sheet_name='Hoja2', index=False)
    return print("Inventario Actualizado")
rewrite_file(inventory_updated, excel_file)